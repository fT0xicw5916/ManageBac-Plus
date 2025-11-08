from enum import StrEnum, Enum
from openpyxl.styles import PatternFill


class ExcelFillColor(Enum):
    GREEN  = PatternFill(start_color="FF9FCE63", end_color="FF9FCE63", fill_type="solid")
    ORANGE = PatternFill(start_color="FFEA9B56", end_color="FFEA9B56", fill_type="solid")
    RED    = PatternFill(start_color="FFEA3323", end_color="FFEA3323", fill_type="solid")


class ExtendedStrEnum(StrEnum):
    @classmethod
    def tolist(cls):
        return [c for c in cls]


class Colors(ExtendedStrEnum):
    BLACK   = "black"
    # BROWN   = "brown"
    RED     = "red"
    # TAN     = "tan"
    ORANGE  = "orange"
    # GOLD    = "gold"
    YELLOW  = "yellow"
    GREEN   = "green"
    CYAN    = "cyan"
    BLUE    = "blue"
    # INDIGO  = "indigo"
    PURPLE  = "purple"
    # MAGENTA = "magenta"
    PINK    = "pink"
